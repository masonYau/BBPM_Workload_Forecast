import pandas as pd
from headers import MasterHeader as mh
from headers import BowHeader as bh
from headers import RamHeader as rh
import logging
import numpy as np
import math
import os

class DataProcessor:

    def __init__(self, current_date: pd.Timestamp, frequency: str):

        self.master_df = pd.DataFrame()
        self.comple_status = {'E2E Completed', 'ReCompleted - WBH'}
        self.wbh_status = {
            'E2E Completed': 'Completed',
            'WBH Imposed': 'WBH',
            'ReCompleted - WBH': 'Completed',
            'Cancelled - All AC Closed': 'Completed',
            'CSEM': 'Completed',
            'Cancelled - KPMG Managed': 'Cancelled',
            'Cancelled - RM Managed': 'Cancelled',
            'IN_PROGRESS': 'WIP',
            'Cancelled - 2nd AC Opening under NTB': 'Cancelled',
            'Cancelled - AC re-opened under NTB': 'Cancelled',
            'Pending BA Approval': 'WIP',
            'Not Loading': 'Not Initiated'
        }
        self.current_date = current_date
        self.frequency = frequency
        self.frequency_days = {"D": 1, "W": 7, "M": 30}
        self.completion_distribution = pd.Series()
        self.input_bow_volume = pd.Series()
        self.bow_volume = pd.Series()
        self.received_volume = pd.Series()
        self.actual_start_volume = pd.Series()
        self.open_received_start_volume = pd.Series()
        self.forecast_completion_volume = pd.Series()
        self.actual_completion_volume = pd.Series()
        self.forecast_wbh_letter_volume = pd.Series()
        self.forecast_wbh_call_volume = pd.Series()
        self.completion_volume = pd.DataFrame()
        self.wbh_letter_volume = pd.DataFrame()
        self.wbh_call_volume = pd.DataFrame()
        self.workload_volume = pd.DataFrame()
        self.output_excel_tables = {}
        self.output_excel_path = None
        self.remaining_bow_volume = pd.DataFrame()
        self.actual_cutoff_date = pd.NaT
        self.remaining_bow_cutoff_date = pd.NaT
        self.actual_start_status = {'Completed', 'WBH', 'Cancelled', 'WIP'}
        self.open_start_status = {'WBH', 'WIP', 'Not Initiated'}
        self.actual_completion_status = {'Completed'}
        self.wbh_letter_days = {
            "PR": 90,
            "Trigger": 60,
        }
        self.wbh_call_days = {
            "PR": 95,
            "Trigger": 65,
        }
        self.output_metric_definitions = [
            {
                "Category": "Start Pipeline",
                "Subcategory": "Plan",
                "Metric": "Planned Start Volume",
                "Display Name": "Planned Start Volume",
                "Source": "Input_BoW_Volume.xlsx",
                "Logic": "Monthly input BoW volume allocated to output periods by calendar-day overlap.",
                "Cutoff": "N/A",
            },
            {
                "Category": "Start Pipeline",
                "Subcategory": "Master File Received",
                "Metric": "Received Start Volume",
                "Display Name": "Received Start Volume",
                "Source": "Master File",
                "Logic": "All received cases aggregated by Original T0 and output period.",
                "Cutoff": "Current date",
            },
            {
                "Category": "Start Pipeline",
                "Subcategory": "Master File WIP",
                "Metric": "WIP Received Start Volume",
                "Display Name": "WIP Received Start Volume",
                "Source": "Master File",
                "Logic": "Received cases still WIP after excluding completed and cancelled status.",
                "Cutoff": "Master File latest Original T0",
            },
            {
                "Category": "Start Pipeline",
                "Subcategory": "Remaining Plan",
                "Metric": "Remaining Planned Start Volume",
                "Display Name": "Remaining Planned Start Volume",
                "Source": "Input_BoW_Volume.xlsx + Master File",
                "Logic": "Input period plan minus received starts, then allocated to output periods after Master File cutoff.",
                "Cutoff": "Master File latest Original T0",
            },
            {
                "Category": "Start Pipeline",
                "Subcategory": "Actual",
                "Metric": "Actual Start Volume",
                "Display Name": "Actual Start Volume",
                "Source": "Master File",
                "Logic": "Cases with status in Completed, WBH, Cancelled, WIP aggregated by Original T0.",
                "Cutoff": "Current date",
            },
            {
                "Category": "Completion",
                "Subcategory": "Actual",
                "Metric": "Actual Completion Volume",
                "Display Name": "Actual Completion Volume",
                "Source": "Master File",
                "Logic": "Completed cases aggregated by Approval/Cancel Date.",
                "Cutoff": "Current date",
            },
            {
                "Category": "Completion",
                "Subcategory": "Forecast",
                "Metric": "Forecast Completion Volume",
                "Display Name": "Forecast Completion Volume",
                "Source": "WIP received starts + remaining planned starts + completion distribution",
                "Logic": "WIP received and remaining planned starts multiplied by case-type completion probability distribution.",
                "Cutoff": "Mixed: Master File cutoff for WIP received, current date for actual completion.",
            },
            {
                "Category": "WBH Action",
                "Subcategory": "Letter",
                "Metric": "Forecast WBH Letter Volume",
                "Display Name": "Forecast WBH Letter Volume",
                "Source": "WIP received starts + remaining planned starts + completion distribution",
                "Logic": "Starts still uncompleted at T+90 for PR or T+60 for Trigger.",
                "Cutoff": "Master File latest Original T0 for WIP received.",
            },
            {
                "Category": "WBH Action",
                "Subcategory": "Call",
                "Metric": "Forecast WBH Call Volume",
                "Display Name": "Forecast WBH Call Volume",
                "Source": "WIP received starts + remaining planned starts + completion distribution",
                "Logic": "Starts still uncompleted at T+95 for PR or T+65 for Trigger.",
                "Cutoff": "Master File latest Original T0 for WIP received.",
            },
        ]

        self.input_completion_percentage_config = {
            "file_path": "Input_Completion_Percentage.xlsx",
            "sheet_names": None,
            "period_column": "Month",
            "percentage_column": "Percentage",
            "frequency": "M",
        }

        self.input_bow_volume_config = {
            "file_path": "Input_BoW_Volume.xlsx",
            "sheet_names": None,
            "period_column": "Month",
            "volume_column": "Volume",
            "frequency": "M",
        }

    def read_data(self):

        tracker_path = "BBPM Case tracker.xlsm"
        if not os.path.exists(tracker_path) and os.path.exists("data/BBPM Case tracker.xlsm"):
            tracker_path = "data/BBPM Case tracker.xlsm"

        master_df = pd.read_excel(tracker_path, sheet_name="Master File", skiprows=1)
        master_df[mh.CIN] = master_df[mh.CIN].astype(str).apply(lambda x: x.lstrip('0'))
        master_df[mh.OriginalT0] = pd.to_datetime(master_df[mh.OriginalT0])
        master_df[mh.ReviewType] = master_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in x else "Trigger")
        self.master_df = master_df



    def calculate_completion_distribution(self, cutoff_days=180):
        sample_st = pd.to_datetime("2026-01-01")
        sample_ed = (self.current_date - pd.Timedelta(
            days=150 + self.current_date.day)).to_period(freq=self.frequency).end_time
        master_df = self.master_df
        period_days = self.frequency_days[self.frequency]

        sample_df = master_df[master_df[mh.OriginalT0].apply(lambda x: sample_st <= x <= sample_ed)]
        sample_df = sample_df.sort_values(by=mh.OriginalT0)
        sample_df[mh.ReviewType] = sample_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in x else "Trigger")
        sample_df['Status'] = sample_df[mh.TaskStatus].map(self.wbh_status)
        sample_df['Case Start'] = master_df[mh.OriginalT0]
        sample_df['Case End'] = sample_df.apply(lambda r: r[mh.ApprovalCancelDate] if r['Status'] == 'Completed' else
                                                r[mh.OriginalT0], axis=1)
        sample_df['Case End'] = sample_df['Case End'].mask(sample_df['Case End'].isna(), master_df[mh.OriginalT0])
        sample_df["N Period"] = (sample_df['Case End'] - sample_df['Case Start']).apply(
            lambda x: x.days // self.frequency_days[self.frequency]
        )
        sample_df = sample_df[sample_df["N Period"].apply(lambda x: 0 <= x <= cutoff_days / period_days - 1)]
        sample_df = sample_df[sample_df["Status"] != "WIP"]

        self.completion_distribution = sample_df.groupby(['Status', "N Period"])[mh.CIN].count() / len(sample_df)

    def read_input_completion_percentage(self):
        config = self.input_completion_percentage_config
        completion_distributions = {}

        for case_type, input_df in self.read_case_type_input_sheets(config):
            input_df = input_df[[config["period_column"], config["percentage_column"]]].dropna()
            input_df[config["period_column"]] = pd.to_numeric(input_df[config["period_column"]], errors="coerce")
            input_df[config["percentage_column"]] = pd.to_numeric(input_df[config["percentage_column"]], errors="coerce")
            input_df = input_df.dropna()

            completion_distribution = input_df.groupby(config["period_column"])[config["percentage_column"]].sum()
            completion_distribution.index = completion_distribution.index.astype(int)
            if len(completion_distribution) > 0 and completion_distribution.max() > 1:
                completion_distribution = completion_distribution / 100

            completion_distribution = self.convert_completion_distribution_frequency(
                completion_distribution,
                input_frequency=config["frequency"]
            )

            for n_period, percentage in completion_distribution.items():
                completion_distributions[(case_type, n_period)] = percentage

        self.completion_distribution = pd.Series(completion_distributions).sort_index()
        if not self.completion_distribution.empty:
            self.completion_distribution.index = pd.MultiIndex.from_tuples(
                self.completion_distribution.index,
                names=["Case Type", "N Period"]
            )
        return self.completion_distribution

    def convert_completion_distribution_frequency(self, completion_distribution: pd.Series, input_frequency: str):
        input_frequency_days = self.frequency_days[input_frequency]
        output_frequency_days = self.frequency_days[self.frequency]

        completion_distribution = completion_distribution.sort_index()
        max_input_period = int(completion_distribution.index.max())
        max_output_period = math.ceil((max_input_period + 1) * input_frequency_days / output_frequency_days) - 1
        output_distribution = pd.Series(0.0, index=range(max_output_period + 1))

        for input_period, percentage in completion_distribution.items():
            input_start_day = int(input_period) * input_frequency_days
            input_end_day = (int(input_period) + 1) * input_frequency_days

            first_output_period = input_start_day // output_frequency_days
            last_output_period = (input_end_day - 1) // output_frequency_days

            for output_period in range(first_output_period, last_output_period + 1):
                output_start_day = output_period * output_frequency_days
                output_end_day = (output_period + 1) * output_frequency_days
                overlap_days = min(input_end_day, output_end_day) - max(input_start_day, output_start_day)
                output_distribution[output_period] += percentage * overlap_days / input_frequency_days

        output_distribution = output_distribution[output_distribution > 0]
        output_distribution.index.name = "N Period"
        return output_distribution

    def read_input_bow_volume(self):
        config = self.input_bow_volume_config
        input_bow_volume = {}
        output_bow_volume = {}

        for case_type, input_df in self.read_case_type_input_sheets(config):
            input_df = input_df[[config["period_column"], config["volume_column"]]].dropna()
            input_df[config["period_column"]] = pd.to_datetime(input_df[config["period_column"]], errors="coerce")
            input_df[config["volume_column"]] = pd.to_numeric(input_df[config["volume_column"]], errors="coerce")
            input_df = input_df.dropna()

            bow_volume = input_df.groupby(config["period_column"])[config["volume_column"]].sum()
            bow_volume.index = pd.to_datetime(bow_volume.index).to_period(config["frequency"])
            bow_volume.index.name = "Input Period"
            bow_volume.name = "Input BoW Volume"

            for input_period, volume in bow_volume.items():
                input_bow_volume[(case_type, input_period)] = volume

            converted_bow_volume = self.convert_bow_volume_frequency(
                bow_volume,
                input_frequency=config["frequency"]
            )

            for period, volume in converted_bow_volume.items():
                output_bow_volume[(case_type, period)] = volume

        self.input_bow_volume = pd.Series(input_bow_volume).sort_index()
        if not self.input_bow_volume.empty:
            self.input_bow_volume.index = pd.MultiIndex.from_tuples(
                self.input_bow_volume.index,
                names=["Case Type", "Input Period"]
            )
        self.input_bow_volume.name = "Input BoW Volume"

        self.bow_volume = pd.Series(output_bow_volume).sort_index()
        if not self.bow_volume.empty:
            self.bow_volume.index = pd.MultiIndex.from_tuples(
                self.bow_volume.index,
                names=["Case Type", "Period"]
            )
        self.bow_volume.name = "BoW Volume"
        return self.bow_volume

    def convert_bow_volume_frequency(self, bow_volume: pd.Series, input_frequency: str):
        output_volume = {}

        for input_period, volume in bow_volume.sort_index().items():
            if not isinstance(input_period, pd.Period):
                input_period = pd.to_datetime(input_period).to_period(input_frequency)
            input_start_date = input_period.start_time.normalize()
            input_end_date = (input_period + 1).start_time.normalize()
            input_days = (input_end_date - input_start_date).days

            first_output_period = input_start_date.to_period(self.frequency)
            last_output_period = (input_end_date - pd.Timedelta(days=1)).to_period(self.frequency)

            for output_period in pd.period_range(first_output_period, last_output_period, freq=self.frequency):
                output_start_date = output_period.start_time.normalize()
                output_end_date = (output_period + 1).start_time.normalize()
                overlap_days = (
                    min(input_end_date, output_end_date) - max(input_start_date, output_start_date)
                ).days

                if overlap_days > 0:
                    output_volume[output_period] = output_volume.get(output_period, 0) + volume * overlap_days / input_days

        output_volume = pd.Series(output_volume).sort_index()
        output_volume.index.name = "Period"
        output_volume.name = "BoW Volume"
        return output_volume

    def resolve_input_file_path(self, file_path):
        if os.path.exists(file_path):
            return file_path

        data_file_path = os.path.join("data", file_path)
        if os.path.exists(data_file_path):
            return data_file_path

        return file_path

    def read_case_type_input_sheets(self, config):
        file_path = self.resolve_input_file_path(config["file_path"])
        excel_file = pd.ExcelFile(file_path)
        sheet_names = config.get("sheet_names")
        if sheet_names is None:
            sheet_names = excel_file.sheet_names
        elif isinstance(sheet_names, str):
            sheet_names = [sheet_names]

        for sheet_name in sheet_names:
            input_df = pd.read_excel(file_path, sheet_name=sheet_name)
            required_columns = {config["period_column"]}
            if "percentage_column" in config:
                required_columns.add(config["percentage_column"])
            if "volume_column" in config:
                required_columns.add(config["volume_column"])

            if not required_columns.issubset(input_df.columns):
                continue

            input_df = input_df.dropna(how="all")
            if input_df.empty:
                continue

            yield sheet_name, input_df

    def get_current_cutoff_date(self):
        return pd.to_datetime(self.current_date).normalize()

    def infer_actual_cutoff_date(self):
        self.actual_cutoff_date = self.get_current_cutoff_date()
        return self.actual_cutoff_date

    def infer_remaining_bow_cutoff_date(self):
        current_date = self.get_current_cutoff_date()
        if mh.OriginalT0 not in self.master_df.columns:
            self.remaining_bow_cutoff_date = current_date
            return self.remaining_bow_cutoff_date

        original_t0 = pd.to_datetime(self.master_df[mh.OriginalT0], errors="coerce").dropna()

        if original_t0.empty:
            self.remaining_bow_cutoff_date = current_date
        else:
            self.remaining_bow_cutoff_date = min(original_t0.max().normalize(), current_date)

        return self.remaining_bow_cutoff_date

    def calculate_received_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if mh.OriginalT0 not in self.master_df.columns or mh.ReviewType not in self.master_df.columns:
            self.received_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
            self.received_volume.name = "Received Volume"
            return self.received_volume

        received_df = self.master_df[[mh.OriginalT0, mh.ReviewType]].copy()
        received_df[mh.ReviewType] = received_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in str(x) else "Trigger")
        received_df[mh.OriginalT0] = pd.to_datetime(
            received_df[mh.OriginalT0],
            errors="coerce"
        ).dt.normalize()
        received_df = received_df.dropna(subset=[mh.OriginalT0])
        received_df = received_df[received_df[mh.OriginalT0] <= cutoff_date]
        received_df["Period"] = received_df[mh.OriginalT0].dt.to_period(self.frequency)

        if received_df.empty:
            self.received_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
        else:
            self.received_volume = received_df.groupby([mh.ReviewType, "Period"]).size()
            self.received_volume.index = self.received_volume.index.set_names(["Case Type", "Period"])
        self.received_volume.name = "Received Volume"
        return self.received_volume

    def calculate_actual_start_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if mh.OriginalT0 not in self.master_df.columns or mh.ReviewType not in self.master_df.columns or mh.TaskStatus not in self.master_df.columns:
            self.actual_start_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
            self.actual_start_volume.name = "Actual Start Volume"
            return self.actual_start_volume

        actual_df = self.master_df[[mh.OriginalT0, mh.ReviewType, mh.TaskStatus]].copy()
        actual_df[mh.OriginalT0] = pd.to_datetime(actual_df[mh.OriginalT0], errors="coerce").dt.normalize()
        actual_df[mh.ReviewType] = actual_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in str(x) else "Trigger")
        actual_df["Status"] = actual_df[mh.TaskStatus].map(self.wbh_status)
        actual_df = actual_df.dropna(subset=[mh.OriginalT0, "Status"])
        actual_df = actual_df[actual_df[mh.OriginalT0] <= cutoff_date]
        actual_df = actual_df[actual_df["Status"].isin(self.actual_start_status)]
        actual_df["Period"] = actual_df[mh.OriginalT0].dt.to_period(self.frequency)

        if actual_df.empty:
            self.actual_start_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
        else:
            self.actual_start_volume = actual_df.groupby([mh.ReviewType, "Period"]).size()
            self.actual_start_volume.index = self.actual_start_volume.index.set_names(["Case Type", "Period"])

        self.actual_start_volume.name = "Actual Start Volume"
        return self.actual_start_volume

    def calculate_open_received_start_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_remaining_bow_cutoff_date()

        if mh.OriginalT0 not in self.master_df.columns or mh.ReviewType not in self.master_df.columns or mh.TaskStatus not in self.master_df.columns:
            self.open_received_start_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Start Period"])
            )
            self.open_received_start_volume.name = "Open Received Start Volume"
            return self.open_received_start_volume

        start_df = self.master_df[[mh.OriginalT0, mh.ReviewType, mh.TaskStatus]].copy()
        start_df[mh.OriginalT0] = pd.to_datetime(start_df[mh.OriginalT0], errors="coerce").dt.normalize()
        start_df[mh.ReviewType] = start_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in str(x) else "Trigger")
        start_df["Status"] = start_df[mh.TaskStatus].map(self.wbh_status)
        start_df = start_df.dropna(subset=[mh.OriginalT0, "Status"])
        start_df = start_df[start_df[mh.OriginalT0] <= cutoff_date]
        start_df = start_df[start_df["Status"].isin(self.open_start_status)]
        start_df["Start Period"] = start_df[mh.OriginalT0].dt.to_period(self.frequency)

        if start_df.empty:
            self.open_received_start_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Start Period"])
            )
        else:
            self.open_received_start_volume = start_df.groupby([mh.ReviewType, "Start Period"]).size().astype(float)
            self.open_received_start_volume.index = self.open_received_start_volume.index.set_names(["Case Type", "Start Period"])

        self.open_received_start_volume.name = "Open Received Start Volume"
        return self.open_received_start_volume

    def calculate_actual_completion_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if mh.ApprovalCancelDate not in self.master_df.columns or mh.ReviewType not in self.master_df.columns or mh.TaskStatus not in self.master_df.columns:
            self.actual_completion_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
            self.actual_completion_volume.name = "Actual Completion Volume"
            return self.actual_completion_volume

        completion_df = self.master_df[[mh.ApprovalCancelDate, mh.ReviewType, mh.TaskStatus]].copy()
        completion_df[mh.ApprovalCancelDate] = pd.to_datetime(
            completion_df[mh.ApprovalCancelDate],
            errors="coerce"
        ).dt.normalize()
        completion_df[mh.ReviewType] = completion_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in str(x) else "Trigger")
        completion_df["Status"] = completion_df[mh.TaskStatus].map(self.wbh_status)
        completion_df = completion_df.dropna(subset=[mh.ApprovalCancelDate, "Status"])
        completion_df = completion_df[completion_df[mh.ApprovalCancelDate] <= cutoff_date]
        completion_df = completion_df[completion_df["Status"].isin(self.actual_completion_status)]
        completion_df["Period"] = completion_df[mh.ApprovalCancelDate].dt.to_period(self.frequency)

        if completion_df.empty:
            self.actual_completion_volume = pd.Series(
                dtype=int,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
        else:
            self.actual_completion_volume = completion_df.groupby([mh.ReviewType, "Period"]).size()
            self.actual_completion_volume.index = self.actual_completion_volume.index.set_names(["Case Type", "Period"])

        self.actual_completion_volume.name = "Actual Completion Volume"
        return self.actual_completion_volume

    def get_case_completion_distribution(self, case_type):
        if self.completion_distribution.empty:
            return pd.Series(dtype=float)

        completion_distribution = self.completion_distribution
        if isinstance(completion_distribution.index, pd.MultiIndex):
            if "Case Type" in completion_distribution.index.names:
                if case_type not in completion_distribution.index.get_level_values("Case Type"):
                    return pd.Series(dtype=float)
                completion_distribution = completion_distribution.xs(case_type, level="Case Type")
            elif "Status" in completion_distribution.index.names:
                if "Completed" not in completion_distribution.index.get_level_values("Status"):
                    return pd.Series(dtype=float)
                completion_distribution = completion_distribution.xs("Completed", level="Status")

        completion_distribution = pd.to_numeric(completion_distribution, errors="coerce").dropna().sort_index()
        completion_distribution.index = completion_distribution.index.astype(int)
        return completion_distribution

    def get_wbh_letter_days(self, case_type):
        normalized_case_type = "PR" if "PR" in str(case_type).upper() else "Trigger"
        return self.wbh_letter_days.get(normalized_case_type)

    def get_wbh_call_days(self, case_type):
        normalized_case_type = "PR" if "PR" in str(case_type).upper() else "Trigger"
        return self.wbh_call_days.get(normalized_case_type)

    def calculate_uncompleted_probability(self, case_type, elapsed_periods):
        completion_distribution = self.get_case_completion_distribution(case_type)
        if completion_distribution.empty:
            return None

        completed_probability = completion_distribution[completion_distribution.index < elapsed_periods].sum()
        return max(1 - completed_probability, 0)

    def add_wbh_letter_volume(self, wbh_letter_volume, case_type, start_period, start_volume, source,
                              condition_on_cutoff=False, cutoff_date=None):
        if start_volume <= 0:
            return

        letter_days = self.get_wbh_letter_days(case_type)
        if letter_days is None:
            return

        letter_days_to_period = letter_days // self.frequency_days[self.frequency]
        letter_period = start_period + letter_days_to_period
        letter_elapsed_periods = max(letter_period.ordinal - start_period.ordinal, 0)
        letter_probability = self.calculate_uncompleted_probability(case_type, letter_elapsed_periods)
        if letter_probability is None or letter_probability <= 0:
            return

        output_period = letter_period
        if condition_on_cutoff and cutoff_date is not None:
            cutoff_period = pd.to_datetime(cutoff_date).to_period(self.frequency)
            cutoff_elapsed_periods = max(cutoff_period.ordinal - start_period.ordinal, 0)

            if letter_elapsed_periods <= cutoff_elapsed_periods:
                letter_probability = 1
            else:
                cutoff_survival_probability = self.calculate_uncompleted_probability(
                    case_type,
                    cutoff_elapsed_periods
                )
                if cutoff_survival_probability is None or cutoff_survival_probability <= 0:
                    return
                letter_probability = letter_probability / cutoff_survival_probability

        wbh_letter_volume[(case_type, output_period, source)] = (
            wbh_letter_volume.get((case_type, output_period, source), 0)
            + start_volume * letter_probability
        )

    def calculate_wbh_letter_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if self.completion_distribution.empty:
            self.read_input_completion_percentage()

        if self.remaining_bow_volume.empty:
            self.calculate_remaining_bow_volume()

        open_received_cutoff_date = self.infer_remaining_bow_cutoff_date()
        open_received_start_volume = self.calculate_open_received_start_volume(cutoff_date=open_received_cutoff_date)
        wbh_letter_volume = {}

        for (case_type, start_period), start_volume in open_received_start_volume.items():
            self.add_wbh_letter_volume(
                wbh_letter_volume,
                case_type,
                start_period,
                start_volume,
                source="Open Received",
                condition_on_cutoff=True,
                cutoff_date=cutoff_date
            )

        if not self.remaining_bow_volume.empty:
            remaining_start_volume = self.remaining_bow_volume["Remaining BoW Volume"]
            for (case_type, start_period), start_volume in remaining_start_volume.items():
                self.add_wbh_letter_volume(
                    wbh_letter_volume,
                    case_type,
                    start_period,
                    start_volume,
                    source="Remaining BoW",
                    condition_on_cutoff=False,
                    cutoff_date=cutoff_date
                )

        if wbh_letter_volume:
            wbh_letter_volume = pd.Series(wbh_letter_volume, dtype=float).sort_index()
            wbh_letter_volume.index = pd.MultiIndex.from_tuples(
                wbh_letter_volume.index,
                names=["Case Type", "Period", "Source"]
            )
        else:
            wbh_letter_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], [], []], names=["Case Type", "Period", "Source"])
            )

        source_volume = wbh_letter_volume.unstack("Source", fill_value=0)
        periods = source_volume.index.sort_values()
        wbh_letter_df = pd.DataFrame(index=periods)
        wbh_letter_df.index = wbh_letter_df.index.set_names(["Case Type", "Period"])
        wbh_letter_df["WIP Received WBH Letter Volume"] = source_volume.get("Open Received", 0)
        wbh_letter_df["Remaining BoW WBH Letter Volume"] = source_volume.get("Remaining BoW", 0)
        wbh_letter_df["Forecast WBH Letter Volume"] = (
            wbh_letter_df["WIP Received WBH Letter Volume"]
            + wbh_letter_df["Remaining BoW WBH Letter Volume"]
        )

        self.forecast_wbh_letter_volume = wbh_letter_df["Forecast WBH Letter Volume"]
        self.forecast_wbh_letter_volume.name = "Forecast WBH Letter Volume"
        self.wbh_letter_volume = wbh_letter_df
        return self.wbh_letter_volume

    def add_wbh_call_volume(self, wbh_call_volume, case_type, start_period, start_volume, source,
                            condition_on_cutoff=False, cutoff_date=None):
        if start_volume <= 0:
            return

        call_days = self.get_wbh_call_days(case_type)
        if call_days is None:
            return

        call_days_to_period = call_days // self.frequency_days[self.frequency]
        call_period = start_period + call_days_to_period
        call_elapsed_periods = max(call_period.ordinal - start_period.ordinal, 0)
        call_probability = self.calculate_uncompleted_probability(case_type, call_elapsed_periods)
        if call_probability is None or call_probability <= 0:
            return

        output_period = call_period
        if condition_on_cutoff and cutoff_date is not None:
            cutoff_period = pd.to_datetime(cutoff_date).to_period(self.frequency)
            cutoff_elapsed_periods = max(cutoff_period.ordinal - start_period.ordinal, 0)

            if call_elapsed_periods <= cutoff_elapsed_periods:
                call_probability = 1
            else:
                cutoff_survival_probability = self.calculate_uncompleted_probability(
                    case_type,
                    cutoff_elapsed_periods
                )
                if cutoff_survival_probability is None or cutoff_survival_probability <= 0:
                    return
                call_probability = call_probability / cutoff_survival_probability

        wbh_call_volume[(case_type, output_period, source)] = (
            wbh_call_volume.get((case_type, output_period, source), 0)
            + start_volume * call_probability
        )

    def calculate_wbh_call_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if self.completion_distribution.empty:
            self.read_input_completion_percentage()

        if self.remaining_bow_volume.empty:
            self.calculate_remaining_bow_volume()

        open_received_cutoff_date = self.infer_remaining_bow_cutoff_date()
        open_received_start_volume = self.calculate_open_received_start_volume(cutoff_date=open_received_cutoff_date)
        wbh_call_volume = {}

        for (case_type, start_period), start_volume in open_received_start_volume.items():
            self.add_wbh_call_volume(
                wbh_call_volume,
                case_type,
                start_period,
                start_volume,
                source="Open Received",
                condition_on_cutoff=True,
                cutoff_date=open_received_cutoff_date
            )

        if not self.remaining_bow_volume.empty:
            remaining_start_volume = self.remaining_bow_volume["Remaining BoW Volume"]
            for (case_type, start_period), start_volume in remaining_start_volume.items():
                self.add_wbh_call_volume(
                    wbh_call_volume,
                    case_type,
                    start_period,
                    start_volume,
                    source="Remaining BoW",
                    condition_on_cutoff=False,
                    cutoff_date=cutoff_date
                )

        if wbh_call_volume:
            wbh_call_volume = pd.Series(wbh_call_volume, dtype=float).sort_index()
            wbh_call_volume.index = pd.MultiIndex.from_tuples(
                wbh_call_volume.index,
                names=["Case Type", "Period", "Source"]
            )
        else:
            wbh_call_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], [], []], names=["Case Type", "Period", "Source"])
            )

        source_volume = wbh_call_volume.unstack("Source", fill_value=0)
        periods = source_volume.index.sort_values()
        wbh_call_df = pd.DataFrame(index=periods)
        wbh_call_df.index = wbh_call_df.index.set_names(["Case Type", "Period"])
        wbh_call_df["WIP Received WBH Call Volume"] = source_volume.get("Open Received", 0)
        wbh_call_df["Remaining BoW WBH Call Volume"] = source_volume.get("Remaining BoW", 0)
        wbh_call_df["Forecast WBH Call Volume"] = (
            wbh_call_df["WIP Received WBH Call Volume"]
            + wbh_call_df["Remaining BoW WBH Call Volume"]
        )

        self.forecast_wbh_call_volume = wbh_call_df["Forecast WBH Call Volume"]
        self.forecast_wbh_call_volume.name = "Forecast WBH Call Volume"
        self.wbh_call_volume = wbh_call_df
        return self.wbh_call_volume

    def add_forecast_completion_volume(self, forecast_completion_volume, case_type, start_period, start_volume,
                                       condition_on_cutoff=False, cutoff_date=None):
        completion_distribution = self.get_case_completion_distribution(case_type)
        if start_volume <= 0 or completion_distribution.empty:
            return

        elapsed_periods = 0
        survival_probability = 1
        if condition_on_cutoff and cutoff_date is not None:
            cutoff_period = pd.to_datetime(cutoff_date).to_period(self.frequency)
            elapsed_periods = max(cutoff_period.ordinal - start_period.ordinal, 0)
            survival_probability = 1 - completion_distribution[completion_distribution.index < elapsed_periods].sum()
            if survival_probability <= 0:
                return

        for n_period, percentage in completion_distribution.items():
            n_period = int(n_period)
            if condition_on_cutoff and n_period < elapsed_periods:
                continue

            completion_period = start_period + n_period
            adjusted_percentage = percentage / survival_probability if condition_on_cutoff else percentage
            forecast_completion_volume[(case_type, completion_period)] = (
                forecast_completion_volume.get((case_type, completion_period), 0)
                + start_volume * adjusted_percentage
            )

    def calculate_completion_volume(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if self.completion_distribution.empty:
            self.read_input_completion_percentage()

        if self.remaining_bow_volume.empty:
            self.calculate_remaining_bow_volume()

        open_received_cutoff_date = self.infer_remaining_bow_cutoff_date()
        open_received_start_volume = self.calculate_open_received_start_volume(cutoff_date=open_received_cutoff_date)
        forecast_completion_volume = {}

        for (case_type, start_period), start_volume in open_received_start_volume.items():
            self.add_forecast_completion_volume(
                forecast_completion_volume,
                case_type,
                start_period,
                start_volume,
                condition_on_cutoff=True,
                cutoff_date=cutoff_date
            )

        if not self.remaining_bow_volume.empty:
            remaining_start_volume = self.remaining_bow_volume["Remaining BoW Volume"]
            for (case_type, start_period), start_volume in remaining_start_volume.items():
                self.add_forecast_completion_volume(
                    forecast_completion_volume,
                    case_type,
                    start_period,
                    start_volume,
                    condition_on_cutoff=False,
                    cutoff_date=cutoff_date
                )

        if forecast_completion_volume:
            self.forecast_completion_volume = pd.Series(forecast_completion_volume, dtype=float).sort_index()
            self.forecast_completion_volume.index = pd.MultiIndex.from_tuples(
                self.forecast_completion_volume.index,
                names=["Case Type", "Period"]
            )
        else:
            self.forecast_completion_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
        self.forecast_completion_volume.name = "Forecast Completion Volume"

        self.calculate_actual_completion_volume(cutoff_date=cutoff_date)

        periods = self.forecast_completion_volume.index.union(self.actual_completion_volume.index).sort_values()
        completion_df = pd.DataFrame(index=periods)
        completion_df.index = completion_df.index.set_names(["Case Type", "Period"])
        completion_df["Forecast Completion Volume"] = self.forecast_completion_volume.reindex(periods, fill_value=0)
        completion_df["Actual Completion Volume"] = self.actual_completion_volume.reindex(periods, fill_value=0)

        self.completion_volume = completion_df
        return self.completion_volume

    def calculate_remaining_bow_volume(self):
        if self.input_bow_volume.empty:
            self.read_input_bow_volume()

        remaining_cutoff_date = self.infer_remaining_bow_cutoff_date()
        self.calculate_received_volume(cutoff_date=remaining_cutoff_date)

        input_frequency = self.input_bow_volume_config["frequency"]
        if mh.OriginalT0 in self.master_df.columns and mh.ReviewType in self.master_df.columns:
            received_df = self.master_df[[mh.OriginalT0, mh.ReviewType]].copy()
            received_df[mh.ReviewType] = received_df[mh.ReviewType].apply(lambda x: "PR" if "PR" in str(x) else "Trigger")
            received_df[mh.OriginalT0] = pd.to_datetime(received_df[mh.OriginalT0], errors="coerce").dt.normalize()
            received_df = received_df.dropna(subset=[mh.OriginalT0])
            received_df = received_df[received_df[mh.OriginalT0] <= remaining_cutoff_date]
            received_df["Input Period"] = received_df[mh.OriginalT0].dt.to_period(input_frequency)
            received_by_input_period = received_df.groupby([mh.ReviewType, "Input Period"]).size()
            received_by_input_period.index = received_by_input_period.index.set_names(["Case Type", "Input Period"])
        else:
            received_by_input_period = pd.Series(dtype=int)

        remaining_output_volume = {}
        remaining_start_limit = remaining_cutoff_date + pd.Timedelta(days=1)

        for (case_type, input_period), planned_volume in self.input_bow_volume.sort_index().items():
            input_start_date = input_period.start_time.normalize()
            input_end_date = (input_period + 1).start_time.normalize()
            received_volume = received_by_input_period.get((case_type, input_period), 0)
            remaining_volume = max(planned_volume - received_volume, 0)
            remaining_start_date = max(input_start_date, remaining_start_limit)

            if remaining_volume <= 0 or remaining_start_date >= input_end_date:
                continue

            remaining_days = (input_end_date - remaining_start_date).days
            first_output_period = remaining_start_date.to_period(self.frequency)
            last_output_period = (input_end_date - pd.Timedelta(days=1)).to_period(self.frequency)

            for output_period in pd.period_range(first_output_period, last_output_period, freq=self.frequency):
                output_start_date = output_period.start_time.normalize()
                output_end_date = (output_period + 1).start_time.normalize()
                overlap_days = (
                    min(input_end_date, output_end_date) - max(remaining_start_date, output_start_date)
                ).days

                if overlap_days > 0:
                    remaining_output_volume[(case_type, output_period)] = (
                        remaining_output_volume.get((case_type, output_period), 0)
                        + remaining_volume * overlap_days / remaining_days
                    )

        if remaining_output_volume:
            remaining_output_volume = pd.Series(remaining_output_volume, dtype=float).sort_index()
            remaining_output_volume.index = pd.MultiIndex.from_tuples(
                remaining_output_volume.index,
                names=["Case Type", "Period"]
            )
        else:
            remaining_output_volume = pd.Series(
                dtype=float,
                index=pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])
            )
        remaining_output_volume.name = "Remaining BoW Volume"

        periods = self.bow_volume.index.union(remaining_output_volume.index).sort_values()
        remaining_df = pd.DataFrame(index=periods)
        remaining_df.index = remaining_df.index.set_names(["Case Type", "Period"])
        remaining_df["Input BoW Volume"] = self.bow_volume.reindex(periods, fill_value=0)
        remaining_df["Received Volume"] = self.received_volume.reindex(periods, fill_value=0)
        remaining_df["Remaining BoW Volume"] = remaining_output_volume.reindex(periods, fill_value=0)
        remaining_df["Over Received Volume"] = (
            remaining_df["Received Volume"] - remaining_df["Input BoW Volume"]
        ).clip(lower=0)

        self.remaining_bow_volume = remaining_df
        return self.remaining_bow_volume


    def calculate_workload(self, cutoff_date=None):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        if self.bow_volume.empty:
            self.read_input_bow_volume()

        if self.remaining_bow_volume.empty:
            self.calculate_remaining_bow_volume()

        received_start_volume = self.calculate_received_volume(cutoff_date=cutoff_date)
        open_received_start_volume = self.calculate_open_received_start_volume(
            cutoff_date=self.infer_remaining_bow_cutoff_date()
        )
        actual_start_volume = self.calculate_actual_start_volume(cutoff_date=cutoff_date)
        completion_volume = self.calculate_completion_volume(cutoff_date=cutoff_date)
        wbh_letter_volume = self.calculate_wbh_letter_volume(cutoff_date=cutoff_date)
        wbh_call_volume = self.calculate_wbh_call_volume(cutoff_date=cutoff_date)

        empty_index = pd.MultiIndex.from_arrays([[], []], names=["Case Type", "Period"])

        def normalize_volume_series(volume_series, name):
            if volume_series is None or volume_series.empty:
                return pd.Series(dtype=float, index=empty_index, name=name)

            volume_series = volume_series.copy()
            volume_series.index = volume_series.index.set_names(["Case Type", "Period"])
            volume_series = pd.to_numeric(volume_series, errors="coerce").fillna(0).astype(float)
            volume_series.name = name
            return volume_series

        remaining_start_volume = pd.Series(dtype=float, index=empty_index)
        if not self.remaining_bow_volume.empty:
            remaining_start_volume = self.remaining_bow_volume["Remaining BoW Volume"]

        workload_series = {
            "Planned Start Volume": normalize_volume_series(self.bow_volume, "Planned Start Volume"),
            "Received Start Volume": normalize_volume_series(received_start_volume, "Received Start Volume"),
            "WIP Received Start Volume": normalize_volume_series(
                open_received_start_volume,
                "WIP Received Start Volume"
            ),
            "Remaining Planned Start Volume": normalize_volume_series(
                remaining_start_volume,
                "Remaining Planned Start Volume"
            ),
            "Actual Start Volume": normalize_volume_series(actual_start_volume, "Actual Start Volume"),
            "Actual Completion Volume": normalize_volume_series(
                completion_volume.get("Actual Completion Volume") if not completion_volume.empty else None,
                "Actual Completion Volume"
            ),
            "Forecast Completion Volume": normalize_volume_series(
                completion_volume.get("Forecast Completion Volume") if not completion_volume.empty else None,
                "Forecast Completion Volume"
            ),
            "Forecast WBH Letter Volume": normalize_volume_series(
                wbh_letter_volume.get("Forecast WBH Letter Volume") if not wbh_letter_volume.empty else None,
                "Forecast WBH Letter Volume"
            ),
            "Forecast WBH Call Volume": normalize_volume_series(
                wbh_call_volume.get("Forecast WBH Call Volume") if not wbh_call_volume.empty else None,
                "Forecast WBH Call Volume"
            ),
        }

        periods = empty_index
        for volume_series in workload_series.values():
            periods = periods.union(volume_series.index)
        periods = periods.sort_values()

        workload_df = pd.DataFrame(index=periods)
        workload_df.index = workload_df.index.set_names(["Case Type", "Period"])
        for column, volume_series in workload_series.items():
            workload_df[column] = volume_series.reindex(periods, fill_value=0)

        self.workload_volume = workload_df
        return self.workload_volume

    def add_period_display_columns(self, output_df, period_columns=None):
        if period_columns is None:
            period_columns = ["Period"]

        output_df = output_df.copy()
        for period_column in period_columns:
            if period_column not in output_df.columns:
                continue

            period_start_column = f"{period_column} Start"
            period_end_column = f"{period_column} End"

            output_df[period_start_column] = output_df[period_column].apply(
                lambda period: period.start_time.normalize() if isinstance(period, pd.Period) else pd.NaT
            )
            output_df[period_end_column] = output_df[period_column].apply(
                lambda period: period.end_time.normalize() if isinstance(period, pd.Period) else pd.NaT
            )
            output_df[period_column] = output_df[period_column].astype(str)

        return output_df

    def output_table_from_series(self, volume_series, value_name, period_columns=None):
        if volume_series is None or volume_series.empty:
            return pd.DataFrame()

        output_df = volume_series.reset_index(name=value_name)
        return self.add_period_display_columns(output_df, period_columns=period_columns)

    def output_table_from_dataframe(self, volume_df, period_columns=None):
        if volume_df is None or volume_df.empty:
            return pd.DataFrame()

        output_df = volume_df.reset_index()
        return self.add_period_display_columns(output_df, period_columns=period_columns)

    def build_completion_distribution_output_table(self):
        distribution_df = self.output_table_from_series(
            self.completion_distribution,
            "Completion Probability",
            period_columns=[]
        )
        if distribution_df.empty:
            return distribution_df

        if {"Case Type", "N Period", "Completion Probability"}.issubset(distribution_df.columns):
            distribution_df = distribution_df.sort_values(["Case Type", "N Period"])
            distribution_df["Cumulative Completion Probability"] = (
                distribution_df.groupby("Case Type")["Completion Probability"].cumsum()
            )
            distribution_df["Remaining Uncompleted Probability"] = (
                1 - distribution_df["Cumulative Completion Probability"]
            ).clip(lower=0)

        return distribution_df

    def write_output_excel(self, output_path=None, output_tables=None):
        if output_path is None:
            output_path = os.path.join("data", "Workload_Forecast_Output.xlsx")

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        if output_tables is None:
            output_tables = self.output_excel_tables

        if not output_tables:
            raise ValueError("No output tables available. Run build_output_excel_tables first.")

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        category_fills = {
            "Start Pipeline": PatternFill("solid", fgColor="E2F0D9"),
            "Completion": PatternFill("solid", fgColor="D9EAF7"),
            "WBH Action": PatternFill("solid", fgColor="FCE4D6"),
        }

        if os.path.exists(output_path):
            try:
                with open(output_path, "a+b"):
                    pass
            except PermissionError:
                base_path, extension = os.path.splitext(output_path)
                timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"{base_path}_{timestamp}{extension}"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, output_df in output_tables.items():
                safe_sheet_name = sheet_name[:31]
                if output_df is None or output_df.empty:
                    output_df = pd.DataFrame({"Message": ["No data"]})
                else:
                    output_df = output_df.copy()

                output_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                worksheet = writer.sheets[safe_sheet_name]
                worksheet.freeze_panes = "F2" if safe_sheet_name == "01_Workload_Wide" else "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                header_by_column = {
                    cell.column: str(cell.value)
                    for cell in worksheet[1]
                    if cell.value is not None
                }

                for row in worksheet.iter_rows(min_row=2):
                    category_value = None
                    for cell in row:
                        if header_by_column.get(cell.column) == "Category":
                            category_value = cell.value
                            break

                    if category_value in category_fills:
                        for cell in row:
                            cell.fill = category_fills[category_value]

                    for cell in row:
                        header = header_by_column.get(cell.column, "")
                        if "Probability" in header:
                            cell.number_format = "0.0%"
                        elif "Volume" in header or header == "Value":
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "#,##0.0"
                        elif "Date" in header or header.endswith("Start") or header.endswith("End"):
                            cell.number_format = "yyyy-mm-dd"
                        elif isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.0"

                for column_cells in worksheet.columns:
                    column_letter = get_column_letter(column_cells[0].column)
                    max_length = 0
                    for cell in column_cells:
                        if cell.value is None:
                            continue
                        max_length = max(max_length, len(str(cell.value)))

                    worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)

        self.output_excel_path = output_path
        return self.output_excel_path

    def build_output_excel_tables(self, cutoff_date=None, output_path=None, write_excel=True):
        if cutoff_date is None:
            cutoff_date = self.infer_actual_cutoff_date()

        workload_df = self.calculate_workload(cutoff_date=cutoff_date)
        metric_definitions = pd.DataFrame(self.output_metric_definitions)
        metric_lookup = metric_definitions.set_index("Metric").to_dict("index")

        control_df = pd.DataFrame(
            [
                {"Item": "Current Date", "Value": self.get_current_cutoff_date()},
                {"Item": "Actual Cutoff Date", "Value": self.infer_actual_cutoff_date()},
                {"Item": "Master File Cutoff Date", "Value": self.infer_remaining_bow_cutoff_date()},
                {"Item": "Output Frequency", "Value": self.frequency},
                {"Item": "Completion Input Frequency", "Value": self.input_completion_percentage_config["frequency"]},
                {"Item": "BoW Input Frequency", "Value": self.input_bow_volume_config["frequency"]},
                {"Item": "WIP Received Status", "Value": ", ".join(sorted(self.open_start_status))},
                {"Item": "Actual Start Status", "Value": ", ".join(sorted(self.actual_start_status))},
                {"Item": "Actual Completion Status", "Value": ", ".join(sorted(self.actual_completion_status))},
                {"Item": "WBH Letter Rule", "Value": "PR T+90, Trigger T+60"},
                {"Item": "WBH Call Rule", "Value": "PR T+95, Trigger T+65"},
            ]
        )

        metric_order = {
            metric: order
            for order, metric in enumerate(metric_definitions["Metric"])
        }
        workload_long_source = workload_df.reset_index().melt(
            id_vars=["Case Type", "Period"],
            var_name="Metric",
            value_name="Volume"
        )
        workload_long_source["Category"] = workload_long_source["Metric"].map(
            lambda metric: metric_lookup.get(metric, {}).get("Category", "Other")
        )
        workload_long_source["Subcategory"] = workload_long_source["Metric"].map(
            lambda metric: metric_lookup.get(metric, {}).get("Subcategory", "Other")
        )
        workload_long_source["Display Name"] = workload_long_source["Metric"].map(
            lambda metric: metric_lookup.get(metric, {}).get("Display Name", metric)
        )
        workload_long_source["Metric Order"] = workload_long_source["Metric"].map(metric_order).fillna(999)

        workload_wide = workload_long_source.pivot_table(
            index=["Category", "Subcategory", "Metric Order", "Metric", "Display Name", "Case Type"],
            columns="Period",
            values="Volume",
            aggfunc="sum",
            fill_value=0
        ).reset_index()
        period_columns = sorted(
            [column for column in workload_wide.columns if isinstance(column, pd.Period)]
        )
        workload_wide = workload_wide[
            ["Category", "Subcategory", "Metric Order", "Metric", "Display Name", "Case Type"]
            + period_columns
        ]
        workload_wide = workload_wide.sort_values(["Metric Order", "Case Type"]).drop(columns=["Metric Order"])
        workload_wide.columns = [
            str(column) if isinstance(column, pd.Period) else column
            for column in workload_wide.columns
        ]

        workload_long = workload_long_source.drop(columns=["Metric Order"])
        workload_long = self.add_period_display_columns(workload_long, period_columns=["Period"])
        workload_long = workload_long[
            [
                "Case Type",
                "Period",
                "Period Start",
                "Period End",
                "Category",
                "Subcategory",
                "Metric",
                "Display Name",
                "Volume",
            ]
        ]

        output_tables = {
            "00_Control": control_df,
            "01_Workload_Wide": workload_wide,
            "02_Workload_Long": workload_long,
            "03_Metric_Definitions": metric_definitions,
            "10_Input_BoW": self.output_table_from_series(
                self.input_bow_volume,
                "Input BoW Volume",
                period_columns=["Input Period"]
            ),
            "11_Planned_Start": self.output_table_from_series(
                self.bow_volume,
                "Planned Start Volume",
                period_columns=["Period"]
            ),
            "12_Start_Reconciliation": self.output_table_from_dataframe(
                self.remaining_bow_volume,
                period_columns=["Period"]
            ),
            "13_WIP_Received_Start": self.output_table_from_series(
                self.open_received_start_volume,
                "WIP Received Start Volume",
                period_columns=["Start Period"]
            ),
            "20_Completion": self.output_table_from_dataframe(
                self.completion_volume,
                period_columns=["Period"]
            ),
            "21_Completion_Distribution": self.build_completion_distribution_output_table(),
            "30_WBH_Letter": self.output_table_from_dataframe(
                self.wbh_letter_volume,
                period_columns=["Period"]
            ),
            "31_WBH_Call": self.output_table_from_dataframe(
                self.wbh_call_volume,
                period_columns=["Period"]
            ),
        }

        self.output_excel_tables = output_tables
        if write_excel or output_path is not None:
            self.write_output_excel(output_path=output_path, output_tables=output_tables)

        return self.output_excel_tables



from datetime import datetime
current_time = pd.to_datetime(datetime.now())
self = DataProcessor(current_time, "M")
self.read_data()
self.calculate_completion_distribution()
self.read_input_completion_percentage()
self.read_input_bow_volume()
self.calculate_remaining_bow_volume()
self.calculate_actual_start_volume()
self.calculate_completion_volume()
self.calculate_wbh_letter_volume()
self.calculate_wbh_call_volume()
self.calculate_workload()
self.build_output_excel_tables()
